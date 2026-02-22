from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import Date, DateTime, MetaData, Table, inspect, select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session
from sqlalchemy.sql.sqltypes import Boolean, Integer, Numeric, Float

from app.api.routers.metadata import LABEL_OVERRIDES
from app.core.config import settings
from app.models.mass_change_batch import MassChangeBatch

_SYSTEM_READONLY_FIELDS = {"created_at", "updated_at"}
_MAX_RETURN_ERRORS = 200


@dataclass
class _ColumnSpec:
    name: str
    column: Any
    is_pk: bool
    is_autoincrement: bool
    nullable: bool
    has_default: bool
    is_readonly: bool
    is_required_create: bool
    is_fk: bool


@dataclass
class _TableSpec:
    table_name: str
    table: Table
    columns: dict[str, _ColumnSpec]
    pk_column_name: str


def _utcnow() -> datetime:
    return datetime.now(UTC)


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")


def _labelize(table_name: str, col_name: str) -> str:
    table_overrides = LABEL_OVERRIDES.get(table_name, {})
    if col_name in table_overrides:
        return table_overrides[col_name]
    return col_name.replace("_", " ").strip().title()


def _serialize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def purge_expired_batches(db: Session) -> int:
    now = _utcnow()
    removed = (
        db.query(MassChangeBatch)
        .filter(MassChangeBatch.expires_at <= now)
        .delete(synchronize_session=False)
    )
    if removed:
        db.commit()
    return int(removed or 0)


def _coerce_bool(raw: Any) -> bool | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return raw
    if isinstance(raw, (int, float)):
        return raw != 0
    text = str(raw).strip().lower()
    if not text:
        return None
    if text in {"true", "1", "yes", "y", "on"}:
        return True
    if text in {"false", "0", "no", "n", "off"}:
        return False
    raise ValueError(f"invalid boolean value '{raw}'")


def _coerce_int(raw: Any) -> int | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return 1 if raw else 0
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        return int(raw)
    text = str(raw).strip()
    if not text:
        return None
    if re.fullmatch(r"-?\d+(\.0+)?", text):
        return int(float(text))
    raise ValueError(f"invalid integer value '{raw}'")


def _coerce_float(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return 1.0 if raw else 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    text = str(raw).strip()
    if not text:
        return None
    return float(text)


def _coerce_date(raw: Any) -> date | None:
    if raw is None:
        return None
    if isinstance(raw, date) and not isinstance(raw, datetime):
        return raw
    if isinstance(raw, datetime):
        return raw.date()
    text = str(raw).strip()
    if not text:
        return None
    return date.fromisoformat(text[:10])


def _coerce_datetime(raw: Any) -> datetime | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw
    if isinstance(raw, date):
        return datetime.combine(raw, time.min)
    text = str(raw).strip()
    if not text:
        return None
    value = text.replace("Z", "+00:00")
    return datetime.fromisoformat(value)


def _canonicalize_fk(value: Any) -> Any:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "|" in text:
        text = text.split("|", 1)[0].strip()
    if re.fullmatch(r"-?\d+(\.0+)?", text):
        return int(float(text))
    return text


def _coerce_for_column(raw: Any, spec: _ColumnSpec) -> Any:
    if raw is None:
        return None
    if isinstance(raw, str) and not raw.strip():
        return None
    if isinstance(spec.column.type, Boolean):
        return _coerce_bool(raw)
    if isinstance(spec.column.type, Integer):
        return _coerce_int(raw)
    if isinstance(spec.column.type, (Float, Numeric)):
        return _coerce_float(raw)
    if isinstance(spec.column.type, Date):
        return _coerce_date(raw)
    if isinstance(spec.column.type, DateTime):
        return _coerce_datetime(raw)
    if spec.is_fk:
        return _canonicalize_fk(raw)
    return str(raw).strip() if isinstance(raw, str) else raw


def _build_table_spec(db: Session, table_name: str) -> _TableSpec:
    bind = db.get_bind()
    insp = inspect(bind)
    if not insp.has_table(table_name):
        raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found.")

    metadata = MetaData()
    table = Table(table_name, metadata, autoload_with=bind)
    pk = insp.get_pk_constraint(table_name) or {}
    pk_cols = list(pk.get("constrained_columns") or [])
    if len(pk_cols) != 1:
        raise HTTPException(
            status_code=400,
            detail=f"Mass submit supports single-column PK only for '{table_name}'.",
        )
    pk_column_name = pk_cols[0]

    fk_columns = set()
    for fk in insp.get_foreign_keys(table_name):
        constrained = fk.get("constrained_columns") or []
        if len(constrained) == 1:
            fk_columns.add(constrained[0])

    columns: dict[str, _ColumnSpec] = {}
    for col in table.columns:
        is_pk = col.name == pk_column_name
        is_autoincrement = bool(getattr(col, "autoincrement", False))
        has_default = col.default is not None or col.server_default is not None
        is_readonly = col.name in _SYSTEM_READONLY_FIELDS
        is_required_create = (
            (not col.nullable)
            and not has_default
            and not is_readonly
            and not (is_pk and is_autoincrement)
        )
        columns[col.name] = _ColumnSpec(
            name=col.name,
            column=col,
            is_pk=is_pk,
            is_autoincrement=is_autoincrement,
            nullable=bool(col.nullable),
            has_default=has_default,
            is_readonly=is_readonly,
            is_required_create=is_required_create,
            is_fk=col.name in fk_columns,
        )

    return _TableSpec(
        table_name=table_name,
        table=table,
        columns=columns,
        pk_column_name=pk_column_name,
    )


def _build_header_map(spec: _TableSpec) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for column_name in spec.columns.keys():
        mapping[_normalize_header(column_name)] = column_name
        mapping[_normalize_header(_labelize(spec.table_name, column_name))] = column_name
    return mapping


def _row_has_values(values: dict[str, Any]) -> bool:
    for value in values.values():
        if value is None:
            continue
        if isinstance(value, str) and not value.strip():
            continue
        return True
    return False


def _pk_exists(db: Session, spec: _TableSpec, pk_value: Any) -> bool:
    pk_col = spec.table.c[spec.pk_column_name]
    stmt = select(pk_col).where(pk_col == pk_value).limit(1)
    return db.execute(stmt).first() is not None


def validate_and_stage_batch(
    db: Session,
    *,
    dataset_key: str,
    table_name: str,
    payload: bytes,
    filename: str,
    user_email: str,
) -> dict[str, Any]:
    purge_expired_batches(db)

    if not payload:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid workbook: {exc}") from exc

    sheet = workbook[workbook.sheetnames[0]] if workbook.sheetnames else None
    if sheet is None:
        raise HTTPException(status_code=400, detail="Workbook has no sheets.")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Workbook has no rows.")

    header_cells = rows[0]
    header_normalized = [_normalize_header(str(cell or "")) for cell in header_cells]
    if not any(header_normalized):
        raise HTTPException(status_code=400, detail="Header row is empty.")

    spec = _build_table_spec(db, table_name)
    header_map = _build_header_map(spec)
    index_to_column: dict[int, str] = {}
    for idx, head in enumerate(header_normalized):
        if not head:
            continue
        column_name = header_map.get(head)
        if column_name:
            index_to_column[idx] = column_name

    if not index_to_column:
        raise HTTPException(
            status_code=400,
            detail="No recognizable table columns found in header row.",
        )

    staged_actions: list[dict[str, Any]] = []
    errors: list[str] = []
    data_rows = 0
    create_rows = 0
    update_rows = 0
    skipped_rows = 0

    for row_idx, excel_row in enumerate(rows[1:], start=2):
        raw_by_column: dict[str, Any] = {}
        for col_idx, col_name in index_to_column.items():
            if col_idx >= len(excel_row):
                continue
            raw_by_column[col_name] = excel_row[col_idx]

        if not _row_has_values(raw_by_column):
            continue
        data_rows += 1

        row_errors: list[str] = []
        parsed_by_column: dict[str, Any] = {}
        for col_name, raw_val in raw_by_column.items():
            spec_col = spec.columns.get(col_name)
            if spec_col is None:
                continue
            try:
                parsed_by_column[col_name] = _coerce_for_column(raw_val, spec_col)
            except Exception as exc:
                row_errors.append(f"Row {row_idx}, field '{col_name}': {exc}")

        if row_errors:
            errors.extend(row_errors)
            continue

        pk_name = spec.pk_column_name
        pk_spec = spec.columns[pk_name]
        pk_value = parsed_by_column.get(pk_name)
        has_pk = pk_value is not None and not (isinstance(pk_value, str) and not pk_value.strip())

        if has_pk:
            if not _pk_exists(db, spec, pk_value):
                errors.append(
                    f"Row {row_idx}: update target not found for {pk_name}={pk_value}."
                )
                continue
            update_payload: dict[str, Any] = {}
            for col_name, value in parsed_by_column.items():
                col_spec = spec.columns[col_name]
                if col_spec.is_pk or col_spec.is_readonly:
                    continue
                if value is None or (isinstance(value, str) and not value.strip()):
                    continue
                update_payload[col_name] = _serialize_value(value)
            if not update_payload:
                skipped_rows += 1
                continue
            staged_actions.append(
                {
                    "row_number": row_idx,
                    "action": "update",
                    "pk": _serialize_value(pk_value),
                    "payload": update_payload,
                }
            )
            update_rows += 1
            continue

        create_payload: dict[str, Any] = {}
        for col_name, value in parsed_by_column.items():
            col_spec = spec.columns[col_name]
            if col_spec.is_readonly:
                continue
            if col_spec.is_pk and col_spec.is_autoincrement:
                continue
            if value is None or (isinstance(value, str) and not value.strip()):
                continue
            create_payload[col_name] = _serialize_value(value)

        missing_required = [
            col_name
            for col_name, col_spec in spec.columns.items()
            if col_spec.is_required_create and col_name not in create_payload
        ]
        if missing_required:
            errors.append(
                f"Row {row_idx}: required fields missing ({', '.join(sorted(missing_required))})."
            )
            continue

        staged_actions.append(
            {
                "row_number": row_idx,
                "action": "create",
                "payload": create_payload,
            }
        )
        create_rows += 1

    now = _utcnow()
    expires_at = now + timedelta(seconds=max(1, settings.MASS_CHANGE_BATCH_TTL_SECONDS))
    batch = MassChangeBatch(
        id=str(uuid4()),
        dataset_key=dataset_key,
        table_name=table_name,
        user_email=(user_email or "system@local").strip().lower(),
        file_name=filename,
        status="validated",
        payload_json=json.dumps(
            {
                "sheet_name": sheet.title,
                "header_columns": len(index_to_column),
                "actions": staged_actions,
            }
        ),
        summary_json=json.dumps(
            {
                "sheet_name": sheet.title,
                "header_columns": len(index_to_column),
                "data_rows": data_rows,
                "staged_rows": len(staged_actions),
                "create_rows": create_rows,
                "update_rows": update_rows,
                "skipped_rows": skipped_rows,
                "errors": len(errors),
            }
        ),
        expires_at=expires_at,
    )
    db.add(batch)
    db.commit()

    return {
        "batch_id": batch.id,
        "dataset_key": dataset_key,
        "file_name": filename,
        "expires_at": expires_at.isoformat(),
        "summary": {
            "sheet_name": sheet.title,
            "header_columns": len(index_to_column),
            "data_rows": data_rows,
            "staged_rows": len(staged_actions),
            "create_rows": create_rows,
            "update_rows": update_rows,
            "skipped_rows": skipped_rows,
            "errors": len(errors),
        },
        "eligible_to_submit": len(staged_actions) > 0,
        "errors": errors[:_MAX_RETURN_ERRORS],
        "warning": (
            None
            if len(errors) <= _MAX_RETURN_ERRORS
            else f"Only first {_MAX_RETURN_ERRORS} errors returned."
        ),
    }


def _get_batch_for_submit(
    db: Session,
    *,
    dataset_key: str,
    batch_id: str,
    user_email: str,
) -> MassChangeBatch:
    purge_expired_batches(db)
    row = (
        db.query(MassChangeBatch)
        .filter(MassChangeBatch.id == batch_id)
        .filter(MassChangeBatch.dataset_key == dataset_key)
        .filter(MassChangeBatch.user_email == (user_email or "system@local").strip().lower())
        .filter(MassChangeBatch.status == "validated")
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Batch not found or no longer valid.")
    if row.expires_at is not None and row.expires_at <= _utcnow():
        db.delete(row)
        db.commit()
        raise HTTPException(status_code=410, detail="Batch expired. Validate again.")
    return row


def submit_staged_batch(
    db: Session,
    *,
    dataset_key: str,
    batch_id: str,
    user_email: str,
) -> dict[str, Any]:
    batch = _get_batch_for_submit(
        db,
        dataset_key=dataset_key,
        batch_id=batch_id,
        user_email=user_email,
    )
    payload = json.loads(batch.payload_json or "{}")
    staged_actions = payload.get("actions") or []
    if not isinstance(staged_actions, list) or not staged_actions:
        raise HTTPException(status_code=400, detail="Batch has no staged rows to submit.")

    spec = _build_table_spec(db, batch.table_name)
    pk_name = spec.pk_column_name

    created = 0
    updated = 0
    failed = 0
    errors: list[str] = []

    for action in staged_actions:
        if not isinstance(action, dict):
            failed += 1
            errors.append("Malformed staged action payload.")
            continue
        row_number = action.get("row_number")
        action_type = str(action.get("action") or "").strip().lower()
        payload_map = action.get("payload") if isinstance(action.get("payload"), dict) else {}

        try:
            coerced_payload: dict[str, Any] = {}
            for col_name, raw_value in payload_map.items():
                spec_col = spec.columns.get(col_name)
                if spec_col is None:
                    continue
                coerced_payload[col_name] = _coerce_for_column(raw_value, spec_col)

            if action_type == "create":
                with db.begin_nested():
                    db.execute(spec.table.insert().values(**coerced_payload))
                created += 1
                continue

            if action_type == "update":
                raw_pk = action.get("pk")
                pk_value = _coerce_for_column(raw_pk, spec.columns[pk_name])
                with db.begin_nested():
                    stmt = (
                        spec.table.update()
                        .where(spec.table.c[pk_name] == pk_value)
                        .values(**coerced_payload)
                    )
                    result = db.execute(stmt)
                    if not result.rowcount:
                        raise ValueError(f"target row not found for {pk_name}={pk_value}")
                updated += 1
                continue

            failed += 1
            errors.append(f"Row {row_number}: unsupported action '{action_type}'.")
        except (IntegrityError, ValueError, TypeError) as exc:
            failed += 1
            errors.append(f"Row {row_number}: {exc}")
        except Exception as exc:
            failed += 1
            errors.append(f"Row {row_number}: submit failed ({exc})")

    batch.status = "submitted"
    batch.submitted_at = _utcnow()
    db.delete(batch)
    db.commit()

    processed = created + updated + failed
    return {
        "message": "Mass submit completed.",
        "summary": {
            "processed": processed,
            "created": created,
            "updated": updated,
            "failed": failed,
        },
        "errors": errors[:_MAX_RETURN_ERRORS],
    }
