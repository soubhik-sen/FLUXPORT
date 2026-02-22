from __future__ import annotations

import json
import re
from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any
from uuid import uuid4

from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import and_, func, or_
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models.customer_master import CustomerMaster
from app.models.document_edit_lock import DocumentEditLock
from app.models.logistics_lookups import (
    PortLookup,
    ShipTypeLookup,
    ShipmentStatusLookup,
    TransportModeLookup,
)
from app.models.mass_change_batch import MassChangeBatch
from app.models.partner_master import PartnerMaster
from app.models.partner_role import PartnerRole
from app.models.po_lookups import (
    PurchaseOrderItemStatusLookup,
    PurchaseOrderStatusLookup,
    PurchaseOrderTypeLookup,
    PurchaseOrgLookup,
)
from app.models.po_schedule_line import POScheduleLine
from app.models.product_master import MaterialMaster
from app.models.purchase_order import PurchaseOrderHeader, PurchaseOrderItem
from app.models.shipment import ShipmentHeader, ShipmentItem
from app.services.role_scope_policy import (
    is_scope_denied,
    resolve_scope_by_field,
    sanitize_scope_by_field,
    scope_deny_detail,
)

_MAX_ISSUES = 400
_DATA_ROWS_LIMIT = 2000
_SUPPLIER_CODES = {"SUPPLIER", "SU"}
_FORWARDER_CODES = {"FORWARDER", "FO", "CARRIER", "CA"}
_MANDATORY_FILL = PatternFill("solid", fgColor="FCE4D6")
_MANDATORY_FONT = Font(bold=True, color="9C0006")

_PO_HEADER_SHEET = "PO_HEADER"
_PO_ITEM_SHEET = "PO_ITEM"
_PO_SCHEDULE_SHEET = "PO_SCHEDULE"
_SHIPMENT_HEADER_SHEET = "SHIPMENT_HEADER"
_SHIPMENT_ITEM_SHEET = "SHIPMENT_ITEM"
_README_SHEET = "README"

_PO_HEADER_COLUMNS = [
    "po_number",
    "type_id",
    "status_id",
    "purchase_org_id",
    "customer_id",
    "vendor_id",
    "forwarder_id",
    "order_date",
    "currency",
]
_PO_HEADER_MANDATORY_CREATE = {
    "po_number",
    "type_id",
    "status_id",
    "purchase_org_id",
    "customer_id",
    "vendor_id",
}
_PO_ITEM_COLUMNS = [
    "po_number",
    "item_number",
    "product_id",
    "status_id",
    "quantity",
    "split_count",
    "unit_price",
    "line_total",
]
_PO_ITEM_MANDATORY_CREATE = {
    "po_number",
    "item_number",
    "product_id",
    "status_id",
    "quantity",
    "split_count",
    "unit_price",
}
_PO_SCHEDULE_COLUMNS = [
    "po_number",
    "item_number",
    "schedule_number",
    "quantity",
    "delivery_date",
]
_PO_SCHEDULE_MANDATORY_CREATE = {
    "po_number",
    "item_number",
    "schedule_number",
    "quantity",
    "delivery_date",
}

_SHIPMENT_HEADER_COLUMNS = [
    "shipment_number",
    "type_id",
    "status_id",
    "mode_id",
    "carrier_id",
    "pol_port_id",
    "pod_port_id",
    "external_reference",
    "master_bill_lading",
    "estimated_departure",
    "estimated_arrival",
    "actual_arrival",
]
_SHIPMENT_HEADER_MANDATORY_CREATE = {
    "shipment_number",
    "type_id",
    "status_id",
    "mode_id",
    "carrier_id",
}
_SHIPMENT_ITEM_COLUMNS = [
    "shipment_number",
    "shipment_item_number",
    "po_number",
    "po_item_number",
    "po_schedule_line_id",
    "shipped_qty",
    "package_id",
    "gross_weight",
]
_SHIPMENT_ITEM_KEY_FIELDS = {
    "shipment_number",
    "shipment_item_number",
    "po_number",
    "po_item_number",
    "po_schedule_line_id",
}
_SHIPMENT_ITEM_MANDATORY_CREATE = set(_SHIPMENT_ITEM_KEY_FIELDS) | {"shipped_qty"}


def _utcnow() -> datetime:
    return datetime.now(UTC)


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (value or "").strip().lower()).strip("_")


def _issue(
    *,
    sheet: str,
    row: int,
    key: str,
    error_code: str,
    reason: str,
    field: str | None = None,
) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "sheet": sheet,
        "row": row,
        "key": key,
        "error_code": error_code,
        "reason": reason,
    }
    if field:
        payload["field"] = field
    return payload


def _key_to_text(key_parts: dict[str, Any]) -> str:
    items = [f"{name}={value}" for name, value in key_parts.items() if value is not None]
    return ", ".join(items) if items else "-"


def _canonicalize_fk(raw: Any) -> Any:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if "|" in text:
        text = text.split("|", 1)[0].strip()
    if re.fullmatch(r"-?\d+(\.0+)?", text):
        return int(float(text))
    return text


def _to_int(raw: Any) -> int | None:
    val = _canonicalize_fk(raw)
    if val is None:
        return None
    if isinstance(val, int):
        return val
    if isinstance(val, str) and re.fullmatch(r"-?\d+", val):
        return int(val)
    raise ValueError(f"invalid integer value '{raw}'")


def _to_float(raw: Any) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, (int, float, Decimal)):
        return float(raw)
    text = str(raw).strip()
    if not text:
        return None
    return float(text)


def _to_date(raw: Any) -> str | None:
    if raw is None:
        return None
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    text = str(raw).strip()
    if not text:
        return None
    return date.fromisoformat(text[:10]).isoformat()


def _parse_iso_date(value: str | None) -> date | None:
    if not value:
        return None
    return date.fromisoformat(value[:10])


def _safe_text(raw: Any) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _bool_filter(scope_by_field: dict[str, set[int]], values_by_field: dict[str, int | None]) -> bool:
    relevant = {
        field: ids
        for field, ids in scope_by_field.items()
        if field in {"customer_id", "company_id", "vendor_id", "forwarder_id"} and ids
    }
    if not relevant:
        return True
    matched = False
    for field, scoped_ids in relevant.items():
        value = values_by_field.get(field)
        if value is None:
            continue
        if value not in scoped_ids:
            return False
        matched = True
    return matched


def _resolve_scope(
    db: Session,
    *,
    user_email: str,
    endpoint_key: str,
    http_method: str,
    endpoint_path: str,
) -> dict[str, set[int]]:
    raw_scope = resolve_scope_by_field(
        db,
        user_email=user_email,
        endpoint_key=endpoint_key,
        http_method=http_method,
        endpoint_path=endpoint_path,
    )
    if is_scope_denied(raw_scope):
        raise HTTPException(status_code=403, detail=scope_deny_detail(raw_scope))
    return sanitize_scope_by_field(raw_scope)


def _append_header(ws, columns: list[str], mandatory: set[str]) -> None:
    ws.append(columns)
    ws.freeze_panes = "A2"
    for idx, column_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=idx)
        if column_name in mandatory:
            cell.fill = _MANDATORY_FILL
            cell.font = _MANDATORY_FONT
        ws.column_dimensions[get_column_letter(idx)].width = max(14, min(36, len(column_name) + 5))


def _add_options_sheet(workbook: Workbook, title_prefix: str, values: list[str]) -> tuple[str, int] | None:
    if not values:
        return None
    sheet_name = f"{title_prefix[:24]}_{len(workbook.sheetnames)}"
    ws = workbook.create_sheet(title=sheet_name)
    ws.append(["Allowed Values"])
    for value in values:
        ws.append([value])
    ws.sheet_state = "hidden"
    return (sheet_name, len(values) + 1)


def _add_dropdown(ws, *, col_idx: int, source_sheet: str, last_row: int) -> None:
    col_letter = get_column_letter(col_idx)
    formula = f"'{source_sheet}'!$A$2:$A${last_row}"
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.promptTitle = "Allowed Value"
    validation.prompt = "Choose an allowed value."
    ws.add_data_validation(validation)
    validation.add(f"{col_letter}2:{col_letter}{_DATA_ROWS_LIMIT}")


def _read_rows(
    workbook,
    *,
    sheet_name: str,
    expected_columns: list[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    if sheet_name not in workbook.sheetnames:
        issues.append(
            _issue(
                sheet=sheet_name,
                row=1,
                key="-",
                error_code="MISSING_SHEET",
                reason=f"Sheet '{sheet_name}' is required.",
            )
        )
        return ([], issues)

    ws = workbook[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        issues.append(
            _issue(
                sheet=sheet_name,
                row=1,
                key="-",
                error_code="EMPTY_SHEET",
                reason=f"Sheet '{sheet_name}' is empty.",
            )
        )
        return ([], issues)

    normalized_headers = [_normalize_header(str(cell or "")) for cell in rows[0]]
    header_pos: dict[str, int] = {name: idx for idx, name in enumerate(normalized_headers) if name}
    for required_col in expected_columns:
        if required_col not in header_pos:
            issues.append(
                _issue(
                    sheet=sheet_name,
                    row=1,
                    key="-",
                    error_code="MISSING_COLUMN",
                    reason=f"Column '{required_col}' is required in sheet header.",
                    field=required_col,
                )
            )

    parsed: list[dict[str, Any]] = []
    for row_index, values in enumerate(rows[1:], start=2):
        row_data: dict[str, Any] = {}
        has_value = False
        for column_name in expected_columns:
            idx = header_pos.get(column_name)
            if idx is None or idx >= len(values):
                row_data[column_name] = None
                continue
            cell_value = values[idx]
            if cell_value is not None and str(cell_value).strip():
                has_value = True
            row_data[column_name] = cell_value
        if not has_value:
            continue
        parsed.append({"row": row_index, "data": row_data})
    return (parsed, issues)


def _po_scope_for_template(db: Session, user_email: str) -> dict[str, set[int]]:
    return _resolve_scope(
        db,
        user_email=user_email,
        endpoint_key="purchase_orders.initialization_data",
        http_method="GET",
        endpoint_path="/api/v1/purchase-orders/initialization-data",
    )


def _po_scope_for_submit(db: Session, user_email: str) -> dict[str, set[int]]:
    return _resolve_scope(
        db,
        user_email=user_email,
        endpoint_key="purchase_orders.create",
        http_method="POST",
        endpoint_path="/api/v1/purchase-orders",
    )


def _shipment_scope_for_template(db: Session, user_email: str) -> dict[str, set[int]]:
    return _resolve_scope(
        db,
        user_email=user_email,
        endpoint_key="shipments.create",
        http_method="POST",
        endpoint_path="/api/v1/shipments",
    )


def _shipment_scope_for_submit(db: Session, user_email: str) -> dict[str, set[int]]:
    return _shipment_scope_for_template(db, user_email)


def _format_opt(value: Any, code: str | None = None, name: str | None = None) -> str:
    right = " | ".join([part for part in [code, name] if part])
    return f"{value} | {right}" if right else str(value)


def _po_template_dropdowns(db: Session, scope: dict[str, set[int]]) -> dict[str, list[str]]:
    type_rows = db.query(PurchaseOrderTypeLookup.id, PurchaseOrderTypeLookup.type_code, PurchaseOrderTypeLookup.type_name).all()
    status_rows = db.query(PurchaseOrderStatusLookup.id, PurchaseOrderStatusLookup.status_code, PurchaseOrderStatusLookup.status_name).all()
    org_rows = db.query(PurchaseOrgLookup.id, PurchaseOrgLookup.org_code, PurchaseOrgLookup.org_name).all()
    item_status_rows = db.query(PurchaseOrderItemStatusLookup.id, PurchaseOrderItemStatusLookup.status_code, PurchaseOrderItemStatusLookup.status_name).all()
    material_rows = db.query(MaterialMaster.id, MaterialMaster.part_number, MaterialMaster.short_description).all()

    customer_query = db.query(CustomerMaster.id, CustomerMaster.customer_identifier, CustomerMaster.legal_name)
    customer_ids = scope.get("customer_id") or set()
    explicit_company_ids = scope.get("company_id") or set()
    if customer_ids:
        customer_query = customer_query.filter(CustomerMaster.id.in_(sorted(customer_ids)))
    elif explicit_company_ids:
        customer_query = customer_query.filter(CustomerMaster.company_id.in_(sorted(explicit_company_ids)))
    customer_rows = customer_query.order_by(CustomerMaster.legal_name.asc()).all()

    vendor_query = (
        db.query(PartnerMaster.id, PartnerMaster.partner_identifier, PartnerMaster.legal_name)
        .join(PartnerRole, PartnerRole.id == PartnerMaster.role_id)
        .filter(func.upper(PartnerRole.role_code).in_(_SUPPLIER_CODES))
    )
    vendor_ids = scope.get("vendor_id") or set()
    if vendor_ids:
        vendor_query = vendor_query.filter(PartnerMaster.id.in_(sorted(vendor_ids)))
    vendor_rows = vendor_query.order_by(PartnerMaster.legal_name.asc()).all()

    forwarder_query = (
        db.query(PartnerMaster.id, PartnerMaster.partner_identifier, PartnerMaster.legal_name)
        .join(PartnerRole, PartnerRole.id == PartnerMaster.role_id)
        .filter(func.upper(PartnerRole.role_code).in_(_FORWARDER_CODES))
    )
    forwarder_ids = scope.get("forwarder_id") or set()
    if forwarder_ids:
        forwarder_query = forwarder_query.filter(PartnerMaster.id.in_(sorted(forwarder_ids)))
    forwarder_rows = forwarder_query.order_by(PartnerMaster.legal_name.asc()).all()

    return {
        "type_id": [_format_opt(r[0], r[1], r[2]) for r in type_rows if r and r[0] is not None],
        "status_id": [_format_opt(r[0], r[1], r[2]) for r in status_rows if r and r[0] is not None],
        "purchase_org_id": [_format_opt(r[0], r[1], r[2]) for r in org_rows if r and r[0] is not None],
        "customer_id": [_format_opt(r[0], r[1], r[2]) for r in customer_rows if r and r[0] is not None],
        "vendor_id": [_format_opt(r[0], r[1], r[2]) for r in vendor_rows if r and r[0] is not None],
        "forwarder_id": [_format_opt(r[0], r[1], r[2]) for r in forwarder_rows if r and r[0] is not None],
        "product_id": [_format_opt(r[0], r[1], r[2]) for r in material_rows if r and r[0] is not None],
        "item_status_id": [_format_opt(r[0], r[1], r[2]) for r in item_status_rows if r and r[0] is not None],
    }


def _shipment_template_dropdowns(db: Session, scope: dict[str, set[int]]) -> dict[str, list[str]]:
    ship_type_rows = db.query(ShipTypeLookup.id, ShipTypeLookup.type_code, ShipTypeLookup.type_name).all()
    ship_status_rows = db.query(ShipmentStatusLookup.id, ShipmentStatusLookup.status_code, ShipmentStatusLookup.status_name).all()
    mode_rows = db.query(TransportModeLookup.id, TransportModeLookup.mode_code, TransportModeLookup.mode_name).all()
    port_rows = db.query(PortLookup.id, PortLookup.port_code, PortLookup.port_name).all()

    carrier_query = (
        db.query(PartnerMaster.id, PartnerMaster.partner_identifier, PartnerMaster.legal_name)
        .join(PartnerRole, PartnerRole.id == PartnerMaster.role_id)
        .filter(func.upper(PartnerRole.role_code).in_(_FORWARDER_CODES))
    )
    forwarder_ids = scope.get("forwarder_id") or set()
    if forwarder_ids:
        carrier_query = carrier_query.filter(PartnerMaster.id.in_(sorted(forwarder_ids)))
    carrier_rows = carrier_query.order_by(PartnerMaster.legal_name.asc()).all()

    return {
        "type_id": [_format_opt(r[0], r[1], r[2]) for r in ship_type_rows if r and r[0] is not None],
        "status_id": [_format_opt(r[0], r[1], r[2]) for r in ship_status_rows if r and r[0] is not None],
        "mode_id": [_format_opt(r[0], r[1], r[2]) for r in mode_rows if r and r[0] is not None],
        "carrier_id": [_format_opt(r[0], r[1], r[2]) for r in carrier_rows if r and r[0] is not None],
        "pol_port_id": [_format_opt(r[0], r[1], r[2]) for r in port_rows if r and r[0] is not None],
        "pod_port_id": [_format_opt(r[0], r[1], r[2]) for r in port_rows if r and r[0] is not None],
    }


def _build_readme_sheet(workbook: Workbook, lines: list[str]) -> None:
    ws = workbook.create_sheet(_README_SHEET)
    ws.append(["Instruction"])
    for line in lines:
        ws.append([line])
    ws.column_dimensions["A"].width = 120


def build_workbook_template(
    db: Session,
    *,
    dataset_key: str,
    user_email: str,
) -> StreamingResponse:
    if dataset_key not in {"po_workbook", "shipment_workbook"}:
        raise HTTPException(status_code=400, detail=f"Unsupported workbook dataset '{dataset_key}'.")

    wb = Workbook()
    wb.remove(wb.active)

    if dataset_key == "po_workbook":
        scope = _po_scope_for_template(db, user_email)
        dropdowns = _po_template_dropdowns(db, scope)

        ws_header = wb.create_sheet(_PO_HEADER_SHEET)
        _append_header(ws_header, _PO_HEADER_COLUMNS, _PO_HEADER_MANDATORY_CREATE)
        ws_item = wb.create_sheet(_PO_ITEM_SHEET)
        _append_header(ws_item, _PO_ITEM_COLUMNS, _PO_ITEM_MANDATORY_CREATE)
        ws_schedule = wb.create_sheet(_PO_SCHEDULE_SHEET)
        _append_header(ws_schedule, _PO_SCHEDULE_COLUMNS, _PO_SCHEDULE_MANDATORY_CREATE)

        po_header_dropdown_fields = {"type_id", "status_id", "purchase_org_id", "customer_id", "vendor_id", "forwarder_id"}
        po_item_dropdown_fields = {"product_id", "status_id"}
        for idx, field in enumerate(_PO_HEADER_COLUMNS, start=1):
            if field not in po_header_dropdown_fields:
                continue
            source_values = dropdowns.get(field) or []
            ref = _add_options_sheet(wb, f"REF_{field}", source_values)
            if ref is not None:
                _add_dropdown(ws_header, col_idx=idx, source_sheet=ref[0], last_row=ref[1])

        for idx, field in enumerate(_PO_ITEM_COLUMNS, start=1):
            if field not in po_item_dropdown_fields:
                continue
            source_key = "item_status_id" if field == "status_id" else field
            source_values = dropdowns.get(source_key) or []
            ref = _add_options_sheet(wb, f"REF_{field}", source_values)
            if ref is not None:
                _add_dropdown(ws_item, col_idx=idx, source_sheet=ref[0], last_row=ref[1])

        _build_readme_sheet(
            wb,
            [
                "PO workbook upsert keys:",
                "PO_HEADER: po_number",
                "PO_ITEM: po_number + item_number",
                "PO_SCHEDULE: po_number + item_number + schedule_number",
                "Rows with existing keys are updated; missing keys are created.",
                "Mandatory columns are highlighted in orange.",
            ],
        )
    else:
        scope = _shipment_scope_for_template(db, user_email)
        dropdowns = _shipment_template_dropdowns(db, scope)

        ws_header = wb.create_sheet(_SHIPMENT_HEADER_SHEET)
        _append_header(ws_header, _SHIPMENT_HEADER_COLUMNS, _SHIPMENT_HEADER_MANDATORY_CREATE)
        ws_item = wb.create_sheet(_SHIPMENT_ITEM_SHEET)
        _append_header(ws_item, _SHIPMENT_ITEM_COLUMNS, _SHIPMENT_ITEM_MANDATORY_CREATE)

        header_dropdown_fields = {"type_id", "status_id", "mode_id", "carrier_id", "pol_port_id", "pod_port_id"}
        for idx, field in enumerate(_SHIPMENT_HEADER_COLUMNS, start=1):
            if field not in header_dropdown_fields:
                continue
            source_values = dropdowns.get(field) or []
            ref = _add_options_sheet(wb, f"REF_{field}", source_values)
            if ref is not None:
                _add_dropdown(ws_header, col_idx=idx, source_sheet=ref[0], last_row=ref[1])

        _build_readme_sheet(
            wb,
            [
                "Shipment workbook upsert keys:",
                "SHIPMENT_HEADER: shipment_number",
                "SHIPMENT_ITEM: shipment_number + shipment_item_number + po_number + po_item_number + po_schedule_line_id",
                "Rows with existing keys are updated; missing keys are created.",
                "Mandatory columns are highlighted in orange.",
                "Transactional foreign-key internals are intentionally not dropdown-driven.",
            ],
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{dataset_key}_template.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


def _batch_payload_result(
    *,
    batch_id: str | None,
    dataset_key: str,
    filename: str,
    expires_at: datetime | None,
    summary: dict[str, Any],
    issues: list[dict[str, Any]],
) -> dict[str, Any]:
    return {
        "batch_id": batch_id or "",
        "dataset_key": dataset_key,
        "file_name": filename,
        "expires_at": expires_at.isoformat() if expires_at is not None else None,
        "summary": summary,
        "eligible_to_submit": bool(batch_id) and not issues,
        "errors": issues[:_MAX_ISSUES],
        "warning": (
            None
            if len(issues) <= _MAX_ISSUES
            else f"Only first {_MAX_ISSUES} issues returned."
        ),
    }


def _save_workbook_batch(
    db: Session,
    *,
    dataset_key: str,
    table_name: str,
    user_email: str,
    filename: str,
    payload: dict[str, Any],
    summary: dict[str, Any],
) -> tuple[str, datetime]:
    now = _utcnow()
    expires_at = now + timedelta(seconds=max(1, settings.MASS_CHANGE_BATCH_TTL_SECONDS))
    batch = MassChangeBatch(
        id=str(uuid4()),
        dataset_key=dataset_key,
        table_name=table_name,
        user_email=(user_email or "system@local").strip().lower(),
        file_name=filename,
        status="validated",
        payload_json=json.dumps(payload),
        summary_json=json.dumps(summary),
        expires_at=expires_at,
    )
    db.add(batch)
    db.commit()
    return (batch.id, expires_at)


def _get_batch_for_submit(
    db: Session,
    *,
    dataset_key: str,
    batch_id: str,
    user_email: str,
) -> MassChangeBatch:
    now = _utcnow()
    row = (
        db.query(MassChangeBatch)
        .filter(MassChangeBatch.id == batch_id)
        .filter(MassChangeBatch.dataset_key == dataset_key)
        .filter(MassChangeBatch.user_email == (user_email or "system@local").strip().lower())
        .filter(MassChangeBatch.status == "validated")
        .first()
    )
    if row is None:
        raise HTTPException(status_code=404, detail="Batch not found or no longer valid.")
    if row.expires_at is not None and row.expires_at <= now:
        db.delete(row)
        db.commit()
        raise HTTPException(status_code=410, detail="Batch expired. Validate again.")
    return row


def _active_lock_issues(
    db: Session,
    *,
    object_type: str,
    document_ids: set[int],
) -> list[dict[str, Any]]:
    if not document_ids:
        return []
    now_naive = datetime.utcnow()
    rows = (
        db.query(DocumentEditLock)
        .filter(DocumentEditLock.object_type == object_type)
        .filter(DocumentEditLock.document_id.in_(sorted(document_ids)))
        .filter(DocumentEditLock.is_active.is_(True))
        .filter(or_(DocumentEditLock.expires_at.is_(None), DocumentEditLock.expires_at > now_naive))
        .all()
    )
    issues: list[dict[str, Any]] = []
    for row in rows:
        issues.append(
            _issue(
                sheet="-",
                row=0,
                key=f"{object_type}:{row.document_id}",
                error_code="LOCK_CONFLICT",
                reason=f"{object_type} {row.document_id} is locked by {row.owner_email}. Submit aborted.",
            )
        )
    return issues


def _validate_po_workbook(
    db: Session,
    *,
    workbook,
    submit_scope: dict[str, set[int]],
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    header_rows, issues = _read_rows(workbook, sheet_name=_PO_HEADER_SHEET, expected_columns=_PO_HEADER_COLUMNS)
    item_rows, item_issues = _read_rows(workbook, sheet_name=_PO_ITEM_SHEET, expected_columns=_PO_ITEM_COLUMNS)
    schedule_rows, sched_issues = _read_rows(workbook, sheet_name=_PO_SCHEDULE_SHEET, expected_columns=_PO_SCHEDULE_COLUMNS)
    issues.extend(item_issues)
    issues.extend(sched_issues)
    if issues:
        return ({}, {"errors": len(issues)}, issues)

    po_numbers = {(_safe_text(r["data"].get("po_number")) or "") for r in header_rows + item_rows + schedule_rows}
    po_numbers = {p for p in po_numbers if p}
    existing_headers = (
        db.query(PurchaseOrderHeader)
        .filter(PurchaseOrderHeader.po_number.in_(sorted(po_numbers)))
        .all()
        if po_numbers
        else []
    )
    header_by_po = {h.po_number: h for h in existing_headers}

    customer_ids: set[int] = set()
    for row in header_rows:
        try:
            cid = _to_int(row["data"].get("customer_id"))
        except Exception:
            cid = None
        if cid is not None:
            customer_ids.add(int(cid))
    customer_rows = (
        db.query(CustomerMaster.id, CustomerMaster.company_id, CustomerMaster.is_active)
        .filter(CustomerMaster.id.in_(sorted(customer_ids)))
        .all()
        if customer_ids
        else []
    )
    customer_map = {
        int(r[0]): {"company_id": r[1], "is_active": bool(r[2])}
        for r in customer_rows
        if r and r[0] is not None
    }

    staged_headers: list[dict[str, Any]] = []
    staged_items: list[dict[str, Any]] = []
    staged_schedules: list[dict[str, Any]] = []
    counts = {"header_create": 0, "header_update": 0, "item_create": 0, "item_update": 0, "schedule_create": 0, "schedule_update": 0}

    for row in header_rows:
        data = row["data"]
        row_no = int(row["row"])
        po_number = _safe_text(data.get("po_number"))
        key_text = _key_to_text({"po_number": po_number})
        if not po_number:
            issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="MISSING_KEY_PART", reason="po_number is required.", field="po_number"))
            continue
        existing = header_by_po.get(po_number)
        mode = "update" if existing is not None else "create"
        try:
            normalized = {
                "po_number": po_number,
                "type_id": _to_int(data.get("type_id")),
                "status_id": _to_int(data.get("status_id")),
                "purchase_org_id": _to_int(data.get("purchase_org_id")),
                "customer_id": _to_int(data.get("customer_id")),
                "vendor_id": _to_int(data.get("vendor_id")),
                "forwarder_id": _to_int(data.get("forwarder_id")),
                "order_date": _to_date(data.get("order_date")),
                "currency": _safe_text(data.get("currency")),
            }
        except Exception as exc:
            issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="TYPE_COERCE_FAILED", reason=str(exc)))
            continue

        if mode == "create":
            missing = [f for f in sorted(_PO_HEADER_MANDATORY_CREATE) if normalized.get(f) in (None, "")]
            if missing:
                issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="MISSING_REQUIRED", reason=f"Missing required fields: {', '.join(missing)}"))
                continue

        resolved_company_id: int | None = None
        cid = normalized.get("customer_id")
        if cid is not None:
            cust = customer_map.get(int(cid))
            if cust is None:
                issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason=f"customer_id={cid} not found.", field="customer_id"))
                continue
            if not cust["is_active"]:
                issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="REF_INACTIVE", reason=f"customer_id={cid} is inactive.", field="customer_id"))
                continue
            if cust["company_id"] is None:
                issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason=f"customer_id={cid} has no company mapping.", field="customer_id"))
                continue
            resolved_company_id = int(cust["company_id"])
        elif existing is not None and existing.company_id is not None:
            resolved_company_id = int(existing.company_id)

        scope_values = {
            "customer_id": int(cid) if cid is not None else (int(existing.customer_id) if existing and existing.customer_id is not None else None),
            "company_id": resolved_company_id if resolved_company_id is not None else (int(existing.company_id) if existing and existing.company_id is not None else None),
            "vendor_id": int(normalized["vendor_id"]) if normalized.get("vendor_id") is not None else (int(existing.vendor_id) if existing and existing.vendor_id is not None else None),
            "forwarder_id": int(normalized["forwarder_id"]) if normalized.get("forwarder_id") is not None else (int(existing.forwarder_id) if existing and existing.forwarder_id is not None else None),
        }
        if not _bool_filter(submit_scope, scope_values):
            issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="SCOPE_DENIED", reason="Row is outside user scope."))
            continue

        staged_headers.append({"row": row_no, "mode": mode, "data": normalized, "resolved_company_id": resolved_company_id})
        counts[f"header_{mode}"] += 1

    staged_header_by_po = {entry["data"]["po_number"]: entry for entry in staged_headers}
    existing_items = (
        db.query(PurchaseOrderItem.id, PurchaseOrderItem.item_number, PurchaseOrderHeader.po_number)
        .join(PurchaseOrderHeader, PurchaseOrderHeader.id == PurchaseOrderItem.po_header_id)
        .filter(PurchaseOrderHeader.po_number.in_(sorted(po_numbers)))
        .all()
        if po_numbers
        else []
    )
    existing_item_map = {(str(row[2]), int(row[1])): int(row[0]) for row in existing_items}

    for row in item_rows:
        data = row["data"]
        row_no = int(row["row"])
        po_number = _safe_text(data.get("po_number"))
        try:
            item_number = _to_int(data.get("item_number"))
        except Exception:
            item_number = None
        key_text = _key_to_text({"po_number": po_number, "item_number": item_number})
        if not po_number or item_number is None:
            issues.append(_issue(sheet=_PO_ITEM_SHEET, row=row_no, key=key_text, error_code="MISSING_KEY_PART", reason="po_number and item_number are required."))
            continue
        if po_number not in header_by_po and po_number not in staged_header_by_po:
            issues.append(_issue(sheet=_PO_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason=f"PO '{po_number}' not found in sheet or database."))
            continue
        mode = "update" if (po_number, int(item_number)) in existing_item_map else "create"
        try:
            normalized = {
                "po_number": po_number,
                "item_number": int(item_number),
                "product_id": _to_int(data.get("product_id")),
                "status_id": _to_int(data.get("status_id")),
                "quantity": _to_float(data.get("quantity")),
                "split_count": _to_int(data.get("split_count")),
                "unit_price": _to_float(data.get("unit_price")),
                "line_total": _to_float(data.get("line_total")),
            }
        except Exception as exc:
            issues.append(_issue(sheet=_PO_ITEM_SHEET, row=row_no, key=key_text, error_code="TYPE_COERCE_FAILED", reason=str(exc)))
            continue
        if normalized.get("line_total") is None and normalized.get("quantity") is not None and normalized.get("unit_price") is not None:
            normalized["line_total"] = round(float(normalized["quantity"]) * float(normalized["unit_price"]), 2)
        if mode == "create":
            missing = [f for f in sorted(_PO_ITEM_MANDATORY_CREATE) if normalized.get(f) in (None, "")]
            if missing:
                issues.append(_issue(sheet=_PO_ITEM_SHEET, row=row_no, key=key_text, error_code="MISSING_REQUIRED", reason=f"Missing required fields: {', '.join(missing)}"))
                continue
        staged_items.append({"row": row_no, "mode": mode, "data": normalized})
        counts[f"item_{mode}"] += 1

    existing_sched_rows = (
        db.query(POScheduleLine.id, POScheduleLine.schedule_number, PurchaseOrderHeader.po_number, PurchaseOrderItem.item_number)
        .join(PurchaseOrderItem, PurchaseOrderItem.id == POScheduleLine.po_item_id)
        .join(PurchaseOrderHeader, PurchaseOrderHeader.id == PurchaseOrderItem.po_header_id)
        .filter(PurchaseOrderHeader.po_number.in_(sorted(po_numbers)))
        .all()
        if po_numbers
        else []
    )
    existing_sched_map = {(str(r[2]), int(r[3]), int(r[1])): int(r[0]) for r in existing_sched_rows}
    sheet_item_keys = {(entry["data"]["po_number"], int(entry["data"]["item_number"])) for entry in staged_items}

    for row in schedule_rows:
        data = row["data"]
        row_no = int(row["row"])
        po_number = _safe_text(data.get("po_number"))
        try:
            item_number = _to_int(data.get("item_number"))
            schedule_number = _to_int(data.get("schedule_number"))
        except Exception:
            item_number = None
            schedule_number = None
        key_text = _key_to_text({"po_number": po_number, "item_number": item_number, "schedule_number": schedule_number})
        if not po_number or item_number is None or schedule_number is None:
            issues.append(_issue(sheet=_PO_SCHEDULE_SHEET, row=row_no, key=key_text, error_code="MISSING_KEY_PART", reason="po_number, item_number, schedule_number are required."))
            continue
        if (po_number, int(item_number)) not in sheet_item_keys and (po_number, int(item_number)) not in existing_item_map:
            issues.append(_issue(sheet=_PO_SCHEDULE_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="Parent PO item not found in sheet or database."))
            continue
        mode = "update" if (po_number, int(item_number), int(schedule_number)) in existing_sched_map else "create"
        try:
            normalized = {
                "po_number": po_number,
                "item_number": int(item_number),
                "schedule_number": int(schedule_number),
                "quantity": _to_float(data.get("quantity")),
                "delivery_date": _to_date(data.get("delivery_date")),
            }
        except Exception as exc:
            issues.append(_issue(sheet=_PO_SCHEDULE_SHEET, row=row_no, key=key_text, error_code="TYPE_COERCE_FAILED", reason=str(exc)))
            continue
        if mode == "create":
            missing = [f for f in sorted(_PO_SCHEDULE_MANDATORY_CREATE) if normalized.get(f) in (None, "")]
            if missing:
                issues.append(_issue(sheet=_PO_SCHEDULE_SHEET, row=row_no, key=key_text, error_code="MISSING_REQUIRED", reason=f"Missing required fields: {', '.join(missing)}"))
                continue
        staged_schedules.append({"row": row_no, "mode": mode, "data": normalized})
        counts[f"schedule_{mode}"] += 1

    summary = {
        "sheet_name": "workbook",
        "data_rows": len(header_rows) + len(item_rows) + len(schedule_rows),
        "header_rows": len(header_rows),
        "item_rows": len(item_rows),
        "schedule_rows": len(schedule_rows),
        "staged_rows": len(staged_headers) + len(staged_items) + len(staged_schedules),
        "create_rows": counts["header_create"] + counts["item_create"] + counts["schedule_create"],
        "update_rows": counts["header_update"] + counts["item_update"] + counts["schedule_update"],
        "errors": len(issues),
        "counts": counts,
    }
    staged_payload = {
        "kind": "po_workbook_v1",
        "headers": staged_headers,
        "items": staged_items,
        "schedules": staged_schedules,
    }
    return (staged_payload, summary, issues)


def _validate_shipment_workbook(
    db: Session,
    *,
    workbook,
    submit_scope: dict[str, set[int]],
) -> tuple[dict[str, Any], dict[str, Any], list[dict[str, Any]]]:
    header_rows, issues = _read_rows(workbook, sheet_name=_SHIPMENT_HEADER_SHEET, expected_columns=_SHIPMENT_HEADER_COLUMNS)
    item_rows, item_issues = _read_rows(workbook, sheet_name=_SHIPMENT_ITEM_SHEET, expected_columns=_SHIPMENT_ITEM_COLUMNS)
    issues.extend(item_issues)
    if issues:
        return ({}, {"errors": len(issues)}, issues)

    shipment_numbers = {(_safe_text(r["data"].get("shipment_number")) or "") for r in header_rows + item_rows}
    shipment_numbers = {s for s in shipment_numbers if s}
    existing_headers = (
        db.query(ShipmentHeader)
        .filter(ShipmentHeader.shipment_number.in_(sorted(shipment_numbers)))
        .all()
        if shipment_numbers
        else []
    )
    header_by_number = {h.shipment_number: h for h in existing_headers}

    po_numbers = {(_safe_text(r["data"].get("po_number")) or "") for r in item_rows}
    po_numbers = {p for p in po_numbers if p}
    po_headers = (
        db.query(PurchaseOrderHeader).filter(PurchaseOrderHeader.po_number.in_(sorted(po_numbers))).all()
        if po_numbers
        else []
    )
    po_header_map = {po.po_number: po for po in po_headers}
    po_items = (
        db.query(PurchaseOrderItem.id, PurchaseOrderItem.item_number, PurchaseOrderHeader.po_number)
        .join(PurchaseOrderHeader, PurchaseOrderHeader.id == PurchaseOrderItem.po_header_id)
        .filter(PurchaseOrderHeader.po_number.in_(sorted(po_numbers)))
        .all()
        if po_numbers
        else []
    )
    po_item_map = {(str(r[2]), int(r[1])): int(r[0]) for r in po_items}
    schedule_ids: set[int] = set()
    for row in item_rows:
        try:
            sid = _to_int(row["data"].get("po_schedule_line_id"))
        except Exception:
            sid = None
        if sid is not None:
            schedule_ids.add(int(sid))
    schedule_rows = (
        db.query(POScheduleLine.id, POScheduleLine.po_item_id)
        .filter(POScheduleLine.id.in_(sorted(schedule_ids)))
        .all()
        if schedule_ids
        else []
    )
    schedule_map = {int(r[0]): int(r[1]) for r in schedule_rows if r and r[0] is not None}
    existing_ship_items = (
        db.query(ShipmentItem.id, ShipmentHeader.shipment_number, ShipmentItem.shipment_item_number, ShipmentItem.po_schedule_line_id, ShipmentItem.po_item_id)
        .join(ShipmentHeader, ShipmentHeader.id == ShipmentItem.shipment_header_id)
        .filter(ShipmentHeader.shipment_number.in_(sorted(shipment_numbers)))
        .all()
        if shipment_numbers
        else []
    )
    existing_item_map = {
        (str(r[1]), int(r[2]), int(r[3]), int(r[4])): int(r[0])
        for r in existing_ship_items
        if r[1] is not None and r[2] is not None and r[3] is not None and r[4] is not None
    }

    staged_headers: list[dict[str, Any]] = []
    staged_items: list[dict[str, Any]] = []
    counts = {"header_create": 0, "header_update": 0, "item_create": 0, "item_update": 0}

    for row in header_rows:
        data = row["data"]
        row_no = int(row["row"])
        shipment_number = _safe_text(data.get("shipment_number"))
        key_text = _key_to_text({"shipment_number": shipment_number})
        if not shipment_number:
            issues.append(_issue(sheet=_SHIPMENT_HEADER_SHEET, row=row_no, key=key_text, error_code="MISSING_KEY_PART", reason="shipment_number is required.", field="shipment_number"))
            continue
        existing = header_by_number.get(shipment_number)
        mode = "update" if existing is not None else "create"
        try:
            normalized = {
                "shipment_number": shipment_number,
                "type_id": _to_int(data.get("type_id")),
                "status_id": _to_int(data.get("status_id")),
                "mode_id": _to_int(data.get("mode_id")),
                "carrier_id": _to_int(data.get("carrier_id")),
                "pol_port_id": _to_int(data.get("pol_port_id")),
                "pod_port_id": _to_int(data.get("pod_port_id")),
                "external_reference": _safe_text(data.get("external_reference")),
                "master_bill_lading": _safe_text(data.get("master_bill_lading")),
                "estimated_departure": _to_date(data.get("estimated_departure")),
                "estimated_arrival": _to_date(data.get("estimated_arrival")),
                "actual_arrival": _to_date(data.get("actual_arrival")),
            }
        except Exception as exc:
            issues.append(_issue(sheet=_SHIPMENT_HEADER_SHEET, row=row_no, key=key_text, error_code="TYPE_COERCE_FAILED", reason=str(exc)))
            continue
        if mode == "create":
            missing = [f for f in sorted(_SHIPMENT_HEADER_MANDATORY_CREATE) if normalized.get(f) in (None, "")]
            if missing:
                issues.append(_issue(sheet=_SHIPMENT_HEADER_SHEET, row=row_no, key=key_text, error_code="MISSING_REQUIRED", reason=f"Missing required fields: {', '.join(missing)}"))
                continue
        forwarder_scope = submit_scope.get("forwarder_id") or set()
        if forwarder_scope:
            candidate = normalized.get("carrier_id")
            if candidate is None and existing is not None and existing.carrier_id is not None:
                candidate = int(existing.carrier_id)
            if candidate is None or int(candidate) not in forwarder_scope:
                issues.append(_issue(sheet=_SHIPMENT_HEADER_SHEET, row=row_no, key=key_text, error_code="SCOPE_DENIED", reason="carrier_id is outside user forwarder scope."))
                continue
        staged_headers.append({"row": row_no, "mode": mode, "data": normalized})
        counts[f"header_{mode}"] += 1

    staged_header_keys = {entry["data"]["shipment_number"] for entry in staged_headers}
    for row in item_rows:
        data = row["data"]
        row_no = int(row["row"])
        shipment_number = _safe_text(data.get("shipment_number"))
        po_number = _safe_text(data.get("po_number"))
        try:
            shipment_item_number = _to_int(data.get("shipment_item_number"))
            po_item_number = _to_int(data.get("po_item_number"))
            po_schedule_line_id = _to_int(data.get("po_schedule_line_id"))
        except Exception:
            shipment_item_number = None
            po_item_number = None
            po_schedule_line_id = None
        key_text = _key_to_text({"shipment_number": shipment_number, "shipment_item_number": shipment_item_number, "po_number": po_number, "po_item_number": po_item_number, "po_schedule_line_id": po_schedule_line_id})
        if not shipment_number or not po_number or shipment_item_number is None or po_item_number is None or po_schedule_line_id is None:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="MISSING_KEY_PART", reason="All shipment item key fields are required."))
            continue
        if shipment_number not in staged_header_keys and shipment_number not in header_by_number:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason=f"Shipment '{shipment_number}' not found in sheet or database."))
            continue
        po_header = po_header_map.get(po_number)
        if po_header is None:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason=f"PO '{po_number}' not found.", field="po_number"))
            continue
        po_item_id = po_item_map.get((po_number, int(po_item_number)))
        if po_item_id is None:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="PO item not found for key."))
            continue
        linked_item = schedule_map.get(int(po_schedule_line_id))
        if linked_item is None:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason=f"po_schedule_line_id={po_schedule_line_id} not found.", field="po_schedule_line_id"))
            continue
        if int(linked_item) != int(po_item_id):
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_MISMATCH", reason="po_schedule_line_id does not belong to po_number + po_item_number."))
            continue
        scope_values = {
            "customer_id": int(po_header.customer_id) if po_header.customer_id is not None else None,
            "company_id": int(po_header.company_id) if po_header.company_id is not None else None,
            "vendor_id": int(po_header.vendor_id) if po_header.vendor_id is not None else None,
            "forwarder_id": int(po_header.forwarder_id) if po_header.forwarder_id is not None else None,
        }
        if not _bool_filter(submit_scope, scope_values):
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="SCOPE_DENIED", reason="Referenced PO context is outside user scope."))
            continue
        mode = "update" if (shipment_number, int(shipment_item_number), int(po_schedule_line_id), int(po_item_id)) in existing_item_map else "create"
        try:
            normalized = {
                "shipment_number": shipment_number,
                "shipment_item_number": int(shipment_item_number),
                "po_number": po_number,
                "po_item_number": int(po_item_number),
                "po_schedule_line_id": int(po_schedule_line_id),
                "po_item_id": int(po_item_id),
                "shipped_qty": _to_float(data.get("shipped_qty")),
                "package_id": _safe_text(data.get("package_id")),
                "gross_weight": _to_float(data.get("gross_weight")),
            }
        except Exception as exc:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="TYPE_COERCE_FAILED", reason=str(exc)))
            continue
        if mode == "create":
            missing = [f for f in sorted(_SHIPMENT_ITEM_MANDATORY_CREATE) if normalized.get(f) in (None, "")]
            if missing:
                issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="MISSING_REQUIRED", reason=f"Missing required fields: {', '.join(missing)}"))
                continue
        staged_items.append({"row": row_no, "mode": mode, "data": normalized})
        counts[f"item_{mode}"] += 1

    summary = {
        "sheet_name": "workbook",
        "data_rows": len(header_rows) + len(item_rows),
        "header_rows": len(header_rows),
        "item_rows": len(item_rows),
        "staged_rows": len(staged_headers) + len(staged_items),
        "create_rows": counts["header_create"] + counts["item_create"],
        "update_rows": counts["header_update"] + counts["item_update"],
        "errors": len(issues),
        "counts": counts,
    }
    staged_payload = {
        "kind": "shipment_workbook_v1",
        "headers": staged_headers,
        "items": staged_items,
    }
    return (staged_payload, summary, issues)


def validate_and_stage_workbook(
    db: Session,
    *,
    dataset_key: str,
    payload: bytes,
    filename: str,
    user_email: str,
    table_name: str,
) -> dict[str, Any]:
    if not payload:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")
    try:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid workbook: {exc}") from exc

    if dataset_key == "po_workbook":
        submit_scope = _po_scope_for_submit(db, user_email)
        staged_payload, summary, issues = _validate_po_workbook(
            db,
            workbook=workbook,
            submit_scope=submit_scope,
        )
    elif dataset_key == "shipment_workbook":
        submit_scope = _shipment_scope_for_submit(db, user_email)
        staged_payload, summary, issues = _validate_shipment_workbook(
            db,
            workbook=workbook,
            submit_scope=submit_scope,
        )
    else:
        raise HTTPException(status_code=400, detail=f"Unsupported workbook dataset '{dataset_key}'.")

    if issues:
        return _batch_payload_result(
            batch_id=None,
            dataset_key=dataset_key,
            filename=filename,
            expires_at=None,
            summary=summary,
            issues=issues,
        )

    staged_rows = int(summary.get("staged_rows") or 0)
    if staged_rows <= 0:
        return _batch_payload_result(
            batch_id=None,
            dataset_key=dataset_key,
            filename=filename,
            expires_at=None,
            summary=summary,
            issues=[
                _issue(
                    sheet="-",
                    row=0,
                    key="-",
                    error_code="NO_DATA_ROWS",
                    reason="No staged rows found in workbook.",
                )
            ],
        )

    batch_payload = {"dataset_key": dataset_key, **staged_payload}
    batch_id, expires_at = _save_workbook_batch(
        db,
        dataset_key=dataset_key,
        table_name=table_name,
        user_email=user_email,
        filename=filename,
        payload=batch_payload,
        summary=summary,
    )
    return _batch_payload_result(
        batch_id=batch_id,
        dataset_key=dataset_key,
        filename=filename,
        expires_at=expires_at,
        summary=summary,
        issues=[],
    )


def _apply_po_workbook_submit(
    db: Session,
    *,
    payload: dict[str, Any],
    submit_scope: dict[str, set[int]],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    headers: list[dict[str, Any]] = list(payload.get("headers") or [])
    items: list[dict[str, Any]] = list(payload.get("items") or [])
    schedules: list[dict[str, Any]] = list(payload.get("schedules") or [])

    po_numbers = {str((entry.get("data") or {}).get("po_number") or "").strip() for entry in headers + items + schedules}
    po_numbers = {p for p in po_numbers if p}
    existing_headers = (
        db.query(PurchaseOrderHeader)
        .filter(PurchaseOrderHeader.po_number.in_(sorted(po_numbers)))
        .all()
        if po_numbers
        else []
    )
    header_by_po = {h.po_number: h for h in existing_headers}
    lock_issues = _active_lock_issues(
        db,
        object_type="PURCHASE_ORDER",
        document_ids={int(h.id) for h in existing_headers},
    )
    if lock_issues:
        return ({}, lock_issues)

    for entry in headers:
        data = dict(entry.get("data") or {})
        row_no = int(entry.get("row") or 0)
        po_number = str(data.get("po_number") or "").strip()
        key_text = _key_to_text({"po_number": po_number})
        if not po_number:
            issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="MISSING_KEY_PART", reason="po_number is required."))
            continue
        header = header_by_po.get(po_number)
        customer_id = data.get("customer_id")
        company_id = entry.get("resolved_company_id")
        if customer_id is not None:
            customer = db.query(CustomerMaster).filter(CustomerMaster.id == int(customer_id)).first()
            if customer is None or not bool(customer.is_active) or customer.company_id is None:
                issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="Invalid customer mapping for row.", field="customer_id"))
                continue
            company_id = int(customer.company_id)

        scope_values = {
            "customer_id": int(customer_id) if customer_id is not None else (int(header.customer_id) if header and header.customer_id is not None else None),
            "company_id": int(company_id) if company_id is not None else (int(header.company_id) if header and header.company_id is not None else None),
            "vendor_id": int(data["vendor_id"]) if data.get("vendor_id") is not None else (int(header.vendor_id) if header and header.vendor_id is not None else None),
            "forwarder_id": int(data["forwarder_id"]) if data.get("forwarder_id") is not None else (int(header.forwarder_id) if header and header.forwarder_id is not None else None),
        }
        if not _bool_filter(submit_scope, scope_values):
            issues.append(_issue(sheet=_PO_HEADER_SHEET, row=row_no, key=key_text, error_code="SCOPE_DENIED", reason="Row is outside user scope."))
            continue

        if header is None:
            header = PurchaseOrderHeader(
                po_number=po_number,
                type_id=int(data["type_id"]),
                status_id=int(data["status_id"]),
                purchase_org_id=int(data["purchase_org_id"]),
                customer_id=int(customer_id) if customer_id is not None else None,
                company_id=int(company_id) if company_id is not None else 0,
                vendor_id=int(data["vendor_id"]),
                forwarder_id=int(data["forwarder_id"]) if data.get("forwarder_id") is not None else None,
                order_date=_parse_iso_date(data.get("order_date")),
                currency=(data.get("currency") or "USD"),
                total_amount=0.0,
            )
            db.add(header)
            db.flush()
            header_by_po[po_number] = header
        else:
            for field in ("type_id", "status_id", "purchase_org_id", "customer_id", "vendor_id", "forwarder_id", "currency"):
                value = data.get(field)
                if value is not None:
                    setattr(header, field, value)
            if company_id is not None:
                header.company_id = int(company_id)
            if data.get("order_date"):
                header.order_date = _parse_iso_date(data.get("order_date"))

    for entry in items:
        data = dict(entry.get("data") or {})
        row_no = int(entry.get("row") or 0)
        po_number = str(data.get("po_number") or "").strip()
        item_number = data.get("item_number")
        key_text = _key_to_text({"po_number": po_number, "item_number": item_number})
        header = header_by_po.get(po_number)
        if header is None:
            issues.append(_issue(sheet=_PO_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="Parent PO header not found."))
            continue
        item = (
            db.query(PurchaseOrderItem)
            .filter(PurchaseOrderItem.po_header_id == int(header.id))
            .filter(PurchaseOrderItem.item_number == int(item_number))
            .first()
        )
        if item is None:
            item = PurchaseOrderItem(
                po_header_id=int(header.id),
                item_number=int(item_number),
                product_id=int(data["product_id"]),
                status_id=int(data["status_id"]),
                quantity=float(data["quantity"]),
                split_count=int(data.get("split_count") or 1),
                unit_price=float(data["unit_price"]),
                line_total=float(data["line_total"] if data.get("line_total") is not None else round(float(data["quantity"]) * float(data["unit_price"]), 2)),
            )
            db.add(item)
        else:
            for field in ("product_id", "status_id", "quantity", "split_count", "unit_price"):
                value = data.get(field)
                if value is not None:
                    setattr(item, field, value)
            if data.get("line_total") is not None:
                item.line_total = float(data["line_total"])

    for entry in schedules:
        data = dict(entry.get("data") or {})
        row_no = int(entry.get("row") or 0)
        po_number = str(data.get("po_number") or "").strip()
        item_number = data.get("item_number")
        schedule_number = data.get("schedule_number")
        key_text = _key_to_text({"po_number": po_number, "item_number": item_number, "schedule_number": schedule_number})
        header = header_by_po.get(po_number)
        if header is None:
            issues.append(_issue(sheet=_PO_SCHEDULE_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="Parent PO header not found."))
            continue
        item = (
            db.query(PurchaseOrderItem)
            .filter(PurchaseOrderItem.po_header_id == int(header.id))
            .filter(PurchaseOrderItem.item_number == int(item_number))
            .first()
        )
        if item is None:
            issues.append(_issue(sheet=_PO_SCHEDULE_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="Parent PO item not found."))
            continue
        schedule = (
            db.query(POScheduleLine)
            .filter(POScheduleLine.po_item_id == int(item.id))
            .filter(POScheduleLine.schedule_number == int(schedule_number))
            .first()
        )
        if schedule is None:
            schedule = POScheduleLine(
                po_item_id=int(item.id),
                schedule_number=int(schedule_number),
                quantity=float(data["quantity"]),
                delivery_date=_parse_iso_date(data["delivery_date"]),
            )
            db.add(schedule)
        else:
            if data.get("quantity") is not None:
                schedule.quantity = float(data["quantity"])
            if data.get("delivery_date"):
                schedule.delivery_date = _parse_iso_date(data["delivery_date"])

    if issues:
        return ({}, issues)
    return (
        {
            "processed": len(headers) + len(items) + len(schedules),
            "headers": len(headers),
            "items": len(items),
            "schedules": len(schedules),
        },
        [],
    )


def _apply_shipment_workbook_submit(
    db: Session,
    *,
    payload: dict[str, Any],
    submit_scope: dict[str, set[int]],
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []
    headers: list[dict[str, Any]] = list(payload.get("headers") or [])
    items: list[dict[str, Any]] = list(payload.get("items") or [])

    shipment_numbers = {str((entry.get("data") or {}).get("shipment_number") or "").strip() for entry in headers + items}
    shipment_numbers = {s for s in shipment_numbers if s}
    existing_headers = (
        db.query(ShipmentHeader).filter(ShipmentHeader.shipment_number.in_(sorted(shipment_numbers))).all()
        if shipment_numbers
        else []
    )
    header_by_number = {h.shipment_number: h for h in existing_headers}
    lock_issues = _active_lock_issues(
        db,
        object_type="SHIPMENT",
        document_ids={int(h.id) for h in existing_headers},
    )
    if lock_issues:
        return ({}, lock_issues)

    for entry in headers:
        data = dict(entry.get("data") or {})
        row_no = int(entry.get("row") or 0)
        shipment_number = str(data.get("shipment_number") or "").strip()
        key_text = _key_to_text({"shipment_number": shipment_number})
        if not shipment_number:
            issues.append(_issue(sheet=_SHIPMENT_HEADER_SHEET, row=row_no, key=key_text, error_code="MISSING_KEY_PART", reason="shipment_number is required."))
            continue
        shipment = header_by_number.get(shipment_number)
        forwarder_scope = submit_scope.get("forwarder_id") or set()
        if forwarder_scope:
            candidate = int(data["carrier_id"]) if data.get("carrier_id") is not None else (int(shipment.carrier_id) if shipment and shipment.carrier_id is not None else None)
            if candidate is None or candidate not in forwarder_scope:
                issues.append(_issue(sheet=_SHIPMENT_HEADER_SHEET, row=row_no, key=key_text, error_code="SCOPE_DENIED", reason="carrier_id is outside user forwarder scope."))
                continue
        if shipment is None:
            shipment = ShipmentHeader(
                shipment_number=shipment_number,
                type_id=int(data["type_id"]),
                status_id=int(data["status_id"]),
                mode_id=int(data["mode_id"]),
                carrier_id=int(data["carrier_id"]),
            )
            if data.get("pol_port_id") is not None:
                shipment.pol_port_id = int(data["pol_port_id"])
            if data.get("pod_port_id") is not None:
                shipment.pod_port_id = int(data["pod_port_id"])
            shipment.external_reference = data.get("external_reference")
            shipment.master_bill_lading = data.get("master_bill_lading")
            shipment.estimated_departure = _parse_iso_date(data.get("estimated_departure"))
            shipment.estimated_arrival = _parse_iso_date(data.get("estimated_arrival"))
            shipment.actual_arrival = _parse_iso_date(data.get("actual_arrival"))
            db.add(shipment)
            db.flush()
            header_by_number[shipment_number] = shipment
        else:
            for field in ("type_id", "status_id", "mode_id", "carrier_id", "pol_port_id", "pod_port_id", "external_reference", "master_bill_lading"):
                value = data.get(field)
                if value is not None:
                    setattr(shipment, field, value)
            if data.get("estimated_departure"):
                shipment.estimated_departure = _parse_iso_date(data.get("estimated_departure"))
            if data.get("estimated_arrival"):
                shipment.estimated_arrival = _parse_iso_date(data.get("estimated_arrival"))
            if data.get("actual_arrival"):
                shipment.actual_arrival = _parse_iso_date(data.get("actual_arrival"))

    for entry in items:
        data = dict(entry.get("data") or {})
        row_no = int(entry.get("row") or 0)
        shipment_number = str(data.get("shipment_number") or "").strip()
        key_text = _key_to_text(
            {
                "shipment_number": shipment_number,
                "shipment_item_number": data.get("shipment_item_number"),
                "po_number": data.get("po_number"),
                "po_item_number": data.get("po_item_number"),
                "po_schedule_line_id": data.get("po_schedule_line_id"),
            }
        )
        shipment = header_by_number.get(shipment_number)
        if shipment is None:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="Parent shipment not found."))
            continue
        po_number = str(data.get("po_number") or "").strip()
        po_item_number = int(data["po_item_number"])
        po_schedule_line_id = int(data["po_schedule_line_id"])
        po_header = db.query(PurchaseOrderHeader).filter(PurchaseOrderHeader.po_number == po_number).first()
        if po_header is None:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason=f"PO '{po_number}' not found."))
            continue
        po_item = (
            db.query(PurchaseOrderItem)
            .filter(PurchaseOrderItem.po_header_id == int(po_header.id))
            .filter(PurchaseOrderItem.item_number == int(po_item_number))
            .first()
        )
        if po_item is None:
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_NOT_FOUND", reason="PO item not found for key."))
            continue
        schedule = db.query(POScheduleLine).filter(POScheduleLine.id == int(po_schedule_line_id)).first()
        if schedule is None or int(schedule.po_item_id) != int(po_item.id):
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="REF_MISMATCH", reason="po_schedule_line_id does not match po_number + po_item_number."))
            continue
        scope_values = {
            "customer_id": int(po_header.customer_id) if po_header.customer_id is not None else None,
            "company_id": int(po_header.company_id) if po_header.company_id is not None else None,
            "vendor_id": int(po_header.vendor_id) if po_header.vendor_id is not None else None,
            "forwarder_id": int(po_header.forwarder_id) if po_header.forwarder_id is not None else None,
        }
        if not _bool_filter(submit_scope, scope_values):
            issues.append(_issue(sheet=_SHIPMENT_ITEM_SHEET, row=row_no, key=key_text, error_code="SCOPE_DENIED", reason="Referenced PO context is outside user scope."))
            continue
        existing = (
            db.query(ShipmentItem)
            .filter(ShipmentItem.shipment_header_id == int(shipment.id))
            .filter(ShipmentItem.shipment_item_number == int(data["shipment_item_number"]))
            .filter(ShipmentItem.po_schedule_line_id == int(po_schedule_line_id))
            .filter(ShipmentItem.po_item_id == int(po_item.id))
            .first()
        )
        if existing is None:
            existing = ShipmentItem(
                shipment_header_id=int(shipment.id),
                po_schedule_line_id=int(po_schedule_line_id),
                po_item_id=int(po_item.id),
                po_number=po_number,
                predecessor_item_no=int(po_item_number),
                shipment_item_number=int(data["shipment_item_number"]),
                shipped_qty=float(data["shipped_qty"]),
                package_id=data.get("package_id"),
                gross_weight=float(data["gross_weight"]) if data.get("gross_weight") is not None else None,
            )
            db.add(existing)
        else:
            if data.get("shipped_qty") is not None:
                existing.shipped_qty = float(data["shipped_qty"])
            if data.get("package_id") is not None:
                existing.package_id = data.get("package_id")
            if data.get("gross_weight") is not None:
                existing.gross_weight = float(data["gross_weight"])

    if issues:
        return ({}, issues)
    return (
        {
            "processed": len(headers) + len(items),
            "headers": len(headers),
            "items": len(items),
        },
        [],
    )


def submit_workbook_batch(
    db: Session,
    *,
    dataset_key: str,
    batch_id: str,
    user_email: str,
) -> dict[str, Any]:
    batch = _get_batch_for_submit(
        db,
        dataset_key=dataset_key,
        batch_id=batch_id,
        user_email=user_email,
    )
    payload = json.loads(batch.payload_json or "{}")
    kind = str(payload.get("kind") or "")
    if dataset_key == "po_workbook" and kind != "po_workbook_v1":
        raise HTTPException(status_code=400, detail="Invalid staged payload kind for PO workbook.")
    if dataset_key == "shipment_workbook" and kind != "shipment_workbook_v1":
        raise HTTPException(status_code=400, detail="Invalid staged payload kind for shipment workbook.")

    try:
        if dataset_key == "po_workbook":
            submit_scope = _po_scope_for_submit(db, user_email)
            summary, issues = _apply_po_workbook_submit(db, payload=payload, submit_scope=submit_scope)
        elif dataset_key == "shipment_workbook":
            submit_scope = _shipment_scope_for_submit(db, user_email)
            summary, issues = _apply_shipment_workbook_submit(db, payload=payload, submit_scope=submit_scope)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported workbook dataset '{dataset_key}'.")

        if issues:
            db.rollback()
            status_code = 409 if any(issue.get("error_code") == "LOCK_CONFLICT" for issue in issues) else 400
            raise HTTPException(
                status_code=status_code,
                detail={
                    "message": "Submit failed. Resolve issues and retry.",
                    "summary": {"processed": 0, "failed": len(issues)},
                    "issues": issues[:_MAX_ISSUES],
                },
            )

        batch.status = "submitted"
        batch.submitted_at = _utcnow()
        db.delete(batch)
        db.commit()
        return {
            "message": "Mass submit completed.",
            "summary": summary,
            "errors": [],
        }
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Submit failed due to unexpected error.",
                "summary": {"processed": 0, "failed": 1},
                "issues": [
                    _issue(
                        sheet="-",
                        row=0,
                        key="-",
                        error_code="SUBMIT_FAILED",
                        reason=str(exc),
                    )
                ],
            },
        ) from exc
