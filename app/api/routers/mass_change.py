from __future__ import annotations

from io import BytesIO
from uuid import uuid4

from fastapi import APIRouter, Body, Depends, HTTPException, Query
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.routers.metadata import download_table_template
from app.core.config import settings
from app.db.session import get_db
from app.services.mass_change_dataset_registry import get_dataset, list_phase1_datasets

router = APIRouter(prefix="/mass-change", tags=["mass-change"])


def _ensure_enabled() -> None:
    if not settings.MASS_CHANGE_ENABLED:
        raise HTTPException(status_code=404, detail="Mass change cockpit is disabled")


def _require_phase1_dataset(dataset_key: str) -> dict:
    row = get_dataset(dataset_key)
    if not row:
        raise HTTPException(status_code=404, detail=f"Unknown dataset '{dataset_key}'")
    if not bool(row.get("phase1_enabled", False)):
        raise HTTPException(
            status_code=403,
            detail=f"Dataset '{dataset_key}' is not enabled in Phase 1.",
        )
    return row


@router.get("/datasets")
def list_datasets():
    _ensure_enabled()
    return {"datasets": list_phase1_datasets()}


@router.get("/{dataset_key}/template.xlsx")
def download_dataset_template(dataset_key: str, db: Session = Depends(get_db)):
    _ensure_enabled()
    dataset = _require_phase1_dataset(dataset_key)
    table_name = str(dataset.get("table_name") or "").strip()
    if not table_name:
        raise HTTPException(status_code=400, detail="Dataset table mapping is missing.")
    return download_table_template(table_name=table_name, db=db)


@router.post("/{dataset_key}/validate")
def validate_dataset_upload(
    dataset_key: str,
    payload: bytes = Body(...),
    filename: str = Query("upload.xlsx"),
):
    _ensure_enabled()
    _require_phase1_dataset(dataset_key)

    filename = (filename or "").strip()
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported.")

    if not payload:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        workbook = load_workbook(filename=BytesIO(payload), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid workbook: {exc}") from exc

    first_sheet = workbook[workbook.sheetnames[0]] if workbook.sheetnames else None
    if first_sheet is None:
        raise HTTPException(status_code=400, detail="Workbook has no sheets.")

    rows = list(first_sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Workbook has no rows.")

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    if not any(header):
        raise HTTPException(status_code=400, detail="Header row is empty.")

    data_rows = [row for row in rows[1:] if any(cell is not None and str(cell).strip() for cell in row)]
    return {
        "batch_id": str(uuid4()),
        "dataset_key": dataset_key,
        "file_name": filename,
        "summary": {
            "sheet_name": first_sheet.title,
            "header_columns": len([h for h in header if h]),
            "data_rows": len(data_rows),
            "errors": 0,
        },
        "eligible_to_submit": True,
        "errors": [],
        "warning": "Phase 1 validate currently performs structure sanity checks only.",
    }


@router.post("/{dataset_key}/submit")
def submit_dataset_upload(dataset_key: str):
    _ensure_enabled()
    _require_phase1_dataset(dataset_key)
    raise HTTPException(
        status_code=501,
        detail="Mass submit engine is not wired yet for this dataset.",
    )
