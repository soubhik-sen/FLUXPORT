from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import cast, column, func, inspect, select, String, table
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.core.decision.attribute_registry import get_attribute_registry

router = APIRouter(prefix="/metadata", tags=["metadata"])


# Optional: per-table/per-column label overrides
LABEL_OVERRIDES: dict[str, dict[str, str]] = {
    "users": {
        "id": "ID",
        "username": "User Name",
        "email": "Email",
        "clearance": "Clearance",
        "is_active": "Active",
    },
    "sys_number_ranges": {
        "doc_category": "Category",
        "doc_type_id": "Type ID",
        "prefix": "Prefix",
        "current_value": "Last Number",
        "padding": "Padding",
        "include_year": "Include Year?",
        "is_active": "Status"
    },
    "sys_workflow_rules": {
        "doc_category": "Category",
        "doc_type_id": "Type ID",
        "state_code": "Current Status",
        "action_key": "Action",
        "required_role_id": "Required Role",
        "is_blocking": "Blocking"
    },
    "currency_lookup": {
        "currency_code": "Currency Code",
        "currency_name": "Currency Name",
        "is_active": "Active"
    },
    "doc_text": {
        "text_type_id": "Text Type",
        "scope_kind": "Scope Kind",
        "po_type_id": "PO Type",
        "ship_type_id": "Shipment Type",
        "document_type_id": "Document Type",
        "customer_id": "Customer",
        "partner_id": "Partner",
        "is_active": "Active",
    },
    "text_val": {
        "doc_text_id": "Doc Text",
        "language": "Language",
        "text_value": "Text Value",
        "valid_from": "Valid From",
        "valid_to": "Valid To",
        "source_type": "Source Type",
        "external_ref": "External Ref",
        "is_active": "Active",
    },
    "event_lookup": {
        "event_code": "Event Code",
        "event_name": "Event Name",
        "event_type": "Event Type",
        "application_object": "Application Object",
        "is_active": "Active",
    },
    "event_profile": {
        "name": "Profile Name",
        "description": "Description",
        "profile_version": "Profile Version",
        "effective_from": "Effective From",
        "effective_to": "Effective To",
        "timezone": "Timezone",
        "created_by": "Created By",
        "last_changed_by": "Last Changed By",
    },
    "profile_event_map": {
        "profile_id": "Event Profile",
        "event_code": "Event Code",
        "inclusion_rule_id": "Inclusion Rule ID",
        "anchor_event_code": "Anchor Event Code",
        "sequence": "Sequence",
        "offset_days": "Offset (dd:hh:mm)",
        "is_mandatory": "Mandatory",
    },
    "event_instance": {
        "parent_id": "Parent ID",
        "po_header_id": "PO Header",
        "shipment_header_id": "Shipment Header",
        "po_number": "PO Number",
        "shipment_number": "Shipment Number",
        "profile_id": "Event Profile",
        "profile_version": "Profile Version",
        "event_code": "Event Code",
        "baseline_date": "Baseline Date",
        "planned_date": "Planned Date",
        "planned_date_manual_override": "Planned Date Manual Override",
        "status_reason": "Status Reason",
        "timezone": "Timezone",
        "actual_date": "Actual Date",
        "status": "Status",
    },
    "text_profile": {
        "name": "Profile Name",
        "object_type": "Object Type",
        "description": "Description",
        "profile_version": "Profile Version",
        "is_active": "Active",
        "effective_from": "Effective From",
        "effective_to": "Effective To",
        "created_by": "Created By",
        "last_changed_by": "Last Changed By",
    },
    "text_profile_rule": {
        "object_type": "Object Type",
        "country_code": "Country",
        "language": "Language",
        "profile_id": "Text Profile",
        "priority": "Priority",
        "is_active": "Active",
        "effective_from": "Effective From",
        "effective_to": "Effective To",
    },
    "profile_text_map": {
        "profile_id": "Text Profile",
        "text_type_id": "Text Type",
        "sequence": "Sequence",
        "is_mandatory": "Mandatory",
        "is_editable": "Editable",
        "is_active": "Active",
    },
    "profile_text_value": {
        "profile_text_map_id": "Profile Text Map",
        "language": "Language",
        "country_code": "Country",
        "text_value": "Text Value",
        "valid_from": "Valid From",
        "valid_to": "Valid To",
        "is_active": "Active",
    },
    "po_text": {
        "po_header_id": "PO Header",
        "profile_id": "Text Profile",
        "profile_version": "Profile Version",
        "text_type_id": "Text Type",
        "language": "Language",
        "text_value": "Text Value",
        "is_user_edited": "User Edited",
    },
    "shipment_text": {
        "shipment_header_id": "Shipment Header",
        "profile_id": "Text Profile",
        "profile_version": "Profile Version",
        "text_type_id": "Text Type",
        "language": "Language",
        "text_value": "Text Value",
        "is_user_edited": "User Edited",
    },
}


@router.get("/attributes/{object_type}")
def get_attribute_metadata(object_type: str) -> dict[str, object]:
    attributes = get_attribute_registry(object_type)
    if not attributes:
        raise HTTPException(status_code=404, detail=f"No attributes found for '{object_type}'.")

    result = [
        {
            "key": key,
            "type": meta.get("type"),
            "label": meta.get("label"),
        }
        for key, meta in attributes.items()
    ]
    return {"objectType": object_type, "attributes": result}


def _labelize(table_name: str, col_name: str) -> str:
    # Use override if present, else snake_case -> Title Case
    table_overrides = LABEL_OVERRIDES.get(table_name, {})
    if col_name in table_overrides:
        return table_overrides[col_name]
    return col_name.replace("_", " ").strip().title()


def _simple_type(sql_type) -> str:
    # Convert SQLAlchemy type to simple string for UI
    t = sql_type.__class__.__name__.lower()

    if "integer" in t or t in ("int", "bigint", "smallint"):
        return "int"
    if "boolean" in t:
        return "bool"
    if "numeric" in t or "decimal" in t:
        return "decimal"
    if "float" in t or "real" in t:
        return "float"
    if "date" == t:
        return "date"
    if "datetime" in t or "timestamp" in t:
        return "datetime"
    if "uuid" in t:
        return "uuid"
    if "json" in t:
        return "json"
    if "text" in t or "string" in t or "varchar" in t or "char" in t:
        return "string"

    # fallback
    return "string"


def _is_searchable(col_name: str, simple_type: str, is_pk: bool) -> bool:
    # Match your example: id not searchable; string/bool searchable by default
    if is_pk:
        return False
    if simple_type in ("string", "bool"):
        return True
    return False


def _build_fk_map(insp, table_name: str) -> dict[str, dict[str, str]]:
    mapping: dict[str, dict[str, str]] = {}
    for fk in insp.get_foreign_keys(table_name):
        constrained_cols = fk.get("constrained_columns") or []
        referred_table = fk.get("referred_table")
        referred_cols = fk.get("referred_columns") or []
        if len(constrained_cols) != 1 or not referred_table or len(referred_cols) != 1:
            continue
        mapping[constrained_cols[0]] = {
            "table": referred_table,
            "column": referred_cols[0],
        }
    return mapping


def _lightweight_table(insp, table_name: str):
    cols = insp.get_columns(table_name)
    return table(table_name, *[column(c["name"]) for c in cols])


def _ensure_table_exists(insp, table_name: str) -> None:
    if not insp.has_table(table_name):
        raise HTTPException(status_code=404, detail=f"Table '{table_name}' not found")


def _pick_display_columns(insp, table_name: str) -> list[str]:
    cols = [c["name"] for c in insp.get_columns(table_name)]
    code_cols = [c for c in cols if c.endswith("_code")]
    name_cols = [c for c in cols if c.endswith("_name")]
    if code_cols and name_cols:
        return [code_cols[0], name_cols[0]]
    if name_cols:
        return [name_cols[0]]
    if "name" in cols:
        return ["name"]
    if code_cols:
        return [code_cols[0]]
    if "legal_name" in cols:
        return ["legal_name"]
    if "trade_name" in cols:
        return ["trade_name"]
    if "description" in cols:
        return ["description"]
    return []


def _fetch_fk_options(
    db: Session,
    insp,
    fk_table: str,
    fk_column: str,
    limit: int = 2000,
) -> list[dict[str, str]]:
    ref_table = _lightweight_table(insp, fk_table)
    if fk_column not in ref_table.c.keys():
        return []

    display_cols = _pick_display_columns(insp, fk_table)
    selectable_cols = [ref_table.c[fk_column]]
    selectable_cols.extend(
        [ref_table.c[c] for c in display_cols if c in ref_table.c.keys() and c != fk_column]
    )

    rows = db.execute(select(*selectable_cols).limit(limit)).mappings().all()
    options: list[dict[str, str]] = []
    for row in rows:
        value = row.get(fk_column)
        if value is None:
            continue
        value_str = str(value)
        text_parts = [str(row.get(c)).strip() for c in display_cols if row.get(c) is not None]
        label = " | ".join([p for p in text_parts if p])
        display = f"{value_str} | {label}" if label else value_str
        options.append({"value": value_str, "display": display})
    return options


@router.get("/{table_name}/fk-options")
def get_table_fk_options(table_name: str, db: Session = Depends(get_db)):
    engine = db.get_bind()
    insp = inspect(engine)

    _ensure_table_exists(insp, table_name)

    fk_map = _build_fk_map(insp, table_name)
    result: dict[str, list[dict[str, str]]] = {}
    for column_name, fk in fk_map.items():
        result[column_name] = _fetch_fk_options(
            db=db,
            insp=insp,
            fk_table=fk["table"],
            fk_column=fk["column"],
        )
    return {"tableName": table_name, "fkOptions": result}


@router.get("/{table_name}/template.xlsx")
def download_table_template(table_name: str, db: Session = Depends(get_db)):
    engine = db.get_bind()
    insp = inspect(engine)

    _ensure_table_exists(insp, table_name)

    cols = insp.get_columns(table_name)
    pk = insp.get_pk_constraint(table_name) or {}
    pk_cols = set(pk.get("constrained_columns") or [])
    fk_map = _build_fk_map(insp, table_name)

    # Template is create-first: exclude PK and system-managed timestamps.
    editable_cols = []
    for c in cols:
        name = c["name"]
        if name in pk_cols:
            continue
        if name in ("created_at", "updated_at"):
            continue
        editable_cols.append(name)

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws.append(editable_cols)
    ws.freeze_panes = "A2"

    fk_col_to_sheet: dict[str, tuple[str, int]] = {}
    for col_name in editable_cols:
        if col_name not in fk_map:
            continue
        fk = fk_map[col_name]
        options = _fetch_fk_options(
            db=db,
            insp=insp,
            fk_table=fk["table"],
            fk_column=fk["column"],
        )
        if not options:
            continue
        ref_sheet_name = f"FK_{col_name[:24]}"
        ref = wb.create_sheet(title=ref_sheet_name)
        ref.append(["Allowed Values"])
        for opt in options:
            ref.append([opt["display"]])
        fk_col_to_sheet[col_name] = (ref_sheet_name, len(options) + 1)

    # Add dropdown validation for FK columns on first 2000 rows.
    for idx, col_name in enumerate(editable_cols, start=1):
        if col_name not in fk_col_to_sheet:
            continue
        sheet_name, last_row = fk_col_to_sheet[col_name]
        col_letter = get_column_letter(idx)
        formula = f"'{sheet_name}'!$A$2:$A${last_row}"
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.promptTitle = "Foreign Key"
        validation.prompt = "Select an allowed value from dropdown."
        ws.add_data_validation(validation)
        validation.add(f"{col_letter}2:{col_letter}2000")

    # Basic widths
    for idx, col_name in enumerate(editable_cols, start=1):
        width = max(14, min(42, len(col_name) + 6))
        ws.column_dimensions[get_column_letter(idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"{table_name}_template.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{table_name}")
def get_table_metadata(table_name: str, db: Session = Depends(get_db)):
    engine = db.get_bind()
    insp = inspect(engine)

    _ensure_table_exists(insp, table_name)

    cols = insp.get_columns(table_name)
    pk = insp.get_pk_constraint(table_name) or {}
    pk_cols = set(pk.get("constrained_columns") or [])
    fk_map = _build_fk_map(insp, table_name)

    result_cols = []
    for c in cols:
        col_name = c["name"]
        stype = _simple_type(c["type"])
        is_pk_col = col_name in pk_cols
        is_system_readonly = col_name in {"version_id", "profile_version"}
        has_default = c.get("default") is not None
        is_autoincrement = bool(c.get("autoincrement"))
        is_required = (
            (not c.get("nullable", True))
            and not is_pk_col
            and not has_default
            and not is_autoincrement
        )

        result_cols.append(
            {
                "key": col_name,
                "label": _labelize(table_name, col_name),
                "type": stype,
                "isSearchable": _is_searchable(col_name, stype, is_pk_col),
                "isReadOnly": is_pk_col or is_system_readonly,
                "isRequired": is_required,
                "isNullable": c.get("nullable", True),
                "isForeignKey": col_name in fk_map,
                "fkTable": fk_map.get(col_name, {}).get("table"),
                "fkColumn": fk_map.get(col_name, {}).get("column"),
            }
        )

    return {"tableName": table_name, "columns": result_cols}


def _to_like_pattern(raw: str) -> str:
    value = raw.strip()
    if not value:
        return ""
    has_wildcard = any(token in value for token in ("*", "?", "%", "_"))
    if has_wildcard:
        return value.replace("*", "%").replace("?", "_")
    return f"%{value}%"


@router.get("/{table_name}/data")
def get_table_data_page(
    table_name: str,
    request: Request,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    engine = db.get_bind()
    insp = inspect(engine)

    _ensure_table_exists(insp, table_name)
    table_obj = _lightweight_table(insp, table_name)
    reserved = {"skip", "limit"}
    filters = []

    for key, raw in request.query_params.items():
        if key in reserved:
            continue
        if key not in table_obj.c.keys():
            continue
        value = (raw or "").strip()
        if not value:
            continue
        pattern = _to_like_pattern(value)
        if not pattern:
            continue
        filters.append(cast(table_obj.c[key], String).ilike(pattern))

    base_stmt = select(table_obj)
    count_stmt = select(func.count()).select_from(table_obj)
    if filters:
        base_stmt = base_stmt.where(*filters)
        count_stmt = count_stmt.where(*filters)

    pk = insp.get_pk_constraint(table_name) or {}
    pk_cols = pk.get("constrained_columns") or []
    if pk_cols and pk_cols[0] in table_obj.c.keys():
        page_stmt = base_stmt.order_by(table_obj.c[pk_cols[0]].desc()).offset(skip).limit(limit)
    else:
        first_col = next(iter(table_obj.c), None)
        if first_col is not None:
            page_stmt = base_stmt.order_by(first_col.asc()).offset(skip).limit(limit)
        else:
            page_stmt = base_stmt.offset(skip).limit(limit)

    rows = db.execute(page_stmt).mappings().all()
    total = int(db.execute(count_stmt).scalar_one())
    return {
        "items": [dict(row) for row in rows],
        "total": total,
        "skip": skip,
        "limit": limit,
    }
